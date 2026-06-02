import tkinter as tk
from tkinter import ttk, messagebox
import openpyxl as op
import os

# Create Excel File
file = "vaccination.xlsx"

if not os.path.exists(file):
    wb = op.Workbook()
    ws = wb.active
    ws.append(["ID","Name","Vaccine"])
    wb.save(file)

# Functions
def display():
    for row in table.get_children():
        table.delete(row)

    wb = op.load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2,values_only=True):
        table.insert("",tk.END,values=row)

def add():
    wb = op.load_workbook(file)
    ws = wb.active

    new_id = ws.max_row

    ws.append([
        new_id,
        name_entry.get(),
        vaccine_entry.get()
    ])

    wb.save(file)

    messagebox.showinfo("Success","Record Added")

    clear()
    display()

def select(event):
    selected = table.focus()

    if selected:
        values = table.item(selected,"values")

        id_var.set(values[0])

        name_entry.delete(0,tk.END)
        name_entry.insert(0,values[1])

        vaccine_entry.delete(0,tk.END)
        vaccine_entry.insert(0,values[2])

def update():
    wb = op.load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == id_var.get():

            row[1].value = name_entry.get()
            row[2].value = vaccine_entry.get()

    wb.save(file)

    messagebox.showinfo("Success","Updated")
    display()

def delete():
    wb = op.load_workbook(file)
    ws = wb.active

    for row in range(2,ws.max_row+1):
        if str(ws.cell(row=1).value) == id_var.get():
            ws.delete_rows(row)
            break

    wb.save(file)

    messagebox.showinfo("Success", "Deleted")

    clear()
    display()

def clear():
    id_var.set("")
    name_entry.delete(0, tk.END)
    vaccine_entry.delete(0, tk.END)

# GUI

Chiko = tk.Tk()
Chiko.title("Vaccination Record System")
Chiko .geometry("700x400")

id_var=tk.StringVar()

tk.Label(Chiko,text="ID").grid(row=0, column=0)
tk.Entry(Chiko,textvariable=id_var, state="readonly").grid(row=0, column=1)

tk.Label(Chiko,text="Name").grid(row=1,column=0)
name_entry = tk.Entry(Chiko)
name_entry.grid(row=1,column=1)

tk.Label(Chiko,text="Vaccine").grid(row=2,column=0)

vaccine_entry = tk.Entry(Chiko)
vaccine_entry.grid(row=2,column=1)

tk.Button(Chiko,text="Add",command=add).grid(row=3,column=0)

tk.Button(Chiko,text="Update",command=update).grid(row=3,column=1)

tk.Button(Chiko,text="Delete",command=delete).grid(row=4,column=0)

tk.Button(Chiko,text="Clear",command=clear).grid(row=4,column=1)

table = ttk.Treeview(Chiko,columns=("ID","Name","Vaccine"),show="headings")

table.heading("ID",text="ID")
table.heading("Name",text="Name")
table.heading("Vaccine",text="Vaccine")

table.grid(row=0,column=3,rowspan=10,padx=20)

table.bind("<ButtonRelease-1>",select)

display()

Chiko.mainloop()
